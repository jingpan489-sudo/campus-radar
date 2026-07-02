"""
投递进度追踪模块
记录你投递的岗位、反馈、进度，生成 Excel 大表。

用法:
    python tracker.py add                # 交互式添加投递记录
    python tracker.py update             # 更新某条记录的状态/反馈
    python tracker.py list               # 查看所有投递记录
    python tracker.py export             # 导出 Excel 表格
    python tracker.py stats              # 查看投递统计
"""
import sqlite3
import os
import sys
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "jobs.db")

# ==================== 投递状态定义 ====================
# 按流程顺序排列，状态会决定 Excel 里的颜色
STATUSES = [
    ("待投递",   "#9E9E9E"),   # 灰色 - 收藏了还没投
    ("已投递",   "#2196F3"),   # 蓝色 - 刚投完
    ("笔试",     "#FF9800"),   # 橙色 - 收到笔试通知
    ("一面",     "#9C27B0"),   # 紫色 - 进入面试
    ("二面",     "#9C27B0"),
    ("三面",     "#9C27B0"),
    ("HR面",     "#00BCD4"),   # 青色 - HR面=快出结果了
    ("Offer",    "#4CAF50"),   # 绿色 - 拿到offer!
    ("已拒",     "#F44336"),   # 红色 - 被拒
    ("已放弃",   "#795548"),   # 棕色 - 自己放弃
]

STATUS_NAMES = [s[0] for s in STATUSES]
STATUS_COLORS = dict(STATUSES)

CHANNELS = ["官网", "内推", "牛客", "邮件", "招聘会", "其他"]


def _get_conn():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS applications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT NOT NULL,         -- 公司名称
            position TEXT NOT NULL,        -- 岗位名称
            apply_date TEXT,               -- 投递日期
            channel TEXT,                  -- 投递渠道
            status TEXT DEFAULT '已投递',   -- 当前状态
            feedback TEXT,                 -- 反馈内容
            next_step TEXT,                -- 下一步动作
            job_url TEXT,                  -- 岗位链接
            update_time TEXT,              -- 最后更新时间
            remark TEXT                    -- 备注
        )
    """)
    return conn


# ==================== 核心操作 ====================

def add_record(company, position, channel="官网", status="已投递",
               feedback="", next_step="", job_url="", remark=""):
    """添加一条投递记录"""
    conn = _get_conn()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    today = date.today().isoformat()
    conn.execute("""
        INSERT INTO applications
        (company, position, apply_date, channel, status, feedback, next_step, job_url, update_time, remark)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (company, position, today, channel, status, feedback, next_step, job_url, now, remark))
    conn.commit()
    conn.close()
    print(f"  ✅ 已添加: {company} - {position} [{status}]")


def update_record(record_id, status=None, feedback=None, next_step=None, remark=None):
    """更新投递记录"""
    conn = _get_conn()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    updates = []
    params = []
    if status:
        updates.append("status = ?")
        params.append(status)
    if feedback is not None:
        updates.append("feedback = ?")
        params.append(feedback)
    if next_step is not None:
        updates.append("next_step = ?")
        params.append(next_step)
    if remark is not None:
        updates.append("remark = ?")
        params.append(remark)
    updates.append("update_time = ?")
    params.append(now)
    params.append(record_id)
    conn.execute(f"UPDATE applications SET {', '.join(updates)} WHERE id = ?", params)
    conn.commit()
    conn.close()
    print(f"  ✅ 已更新记录 #{record_id}")


def list_records(status_filter=None):
    """查看所有投递记录"""
    conn = _get_conn()
    if status_filter:
        rows = conn.execute(
            "SELECT * FROM applications WHERE status=? ORDER BY apply_date DESC",
            (status_filter,)).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM applications ORDER BY apply_date DESC").fetchall()
    conn.close()
    _print_records(rows)
    return rows


def get_stats():
    """投递统计"""
    conn = _get_conn()
    total = conn.execute("SELECT COUNT(*) FROM applications").fetchone()[0]
    by_status = conn.execute(
        "SELECT status, COUNT(*) FROM applications GROUP BY status ORDER BY COUNT(*) DESC"
    ).fetchall()
    by_company = conn.execute(
        "SELECT company, COUNT(*) FROM applications GROUP BY company ORDER BY COUNT(*) DESC LIMIT 10"
    ).fetchall()
    conn.close()

    print(f"\n{'='*50}")
    print(f"📊 投递统计 ｜ 共投递 {total} 个岗位")
    print(f"{'='*50}")
    print("\n按状态分布:")
    for s, n in by_status:
        bar = "█" * n
        print(f"  {s:6s} {n:3d} {bar}")
    if by_company:
        print("\n投递最多的公司:")
        for c, n in by_company:
            print(f"  {c:12s} {n}")
    print()
    return {"total": total, "by_status": dict(by_status)}


def export_excel():
    """导出 Excel 大表，带颜色样式"""
    conn = _get_conn()
    rows = conn.execute("""
        SELECT id, company, position, apply_date, channel, status,
               feedback, next_step, job_url, update_time, remark
        FROM applications ORDER BY apply_date DESC, update_time DESC
    """).fetchall()
    conn.close()

    wb = Workbook()

    # ========== Sheet 1: 投递总表 ==========
    ws = wb.active
    ws.title = "投递总表"

    headers = ["序号", "公司", "岗位", "投递日期", "渠道",
               "状态", "反馈", "下一步", "岗位链接", "更新时间", "备注"]
    # 表头样式
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 数据行
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val else "")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            # 状态列着色
            if col_idx == 6 and val:
                hex_color = STATUS_COLORS.get(val, "#FFFFFF")[1:]
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

    # 列宽
    col_widths = [6, 14, 28, 12, 8, 10, 35, 20, 30, 18, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"  # 冻结表头

    # ========== Sheet 2: 统计看板 ==========
    ws2 = wb.create_sheet("统计看板")
    conn = _get_conn()
    total = conn.execute("SELECT COUNT(*) FROM applications").fetchone()[0]
    by_status = conn.execute(
        "SELECT status, COUNT(*) FROM applications GROUP BY status").fetchall()
    by_channel = conn.execute(
        "SELECT channel, COUNT(*) FROM applications GROUP BY channel").fetchall()
    by_company = conn.execute(
        "SELECT company, COUNT(*) FROM applications GROUP BY company ORDER BY COUNT(*) DESC"
    ).fetchall()
    conn.close()

    ws2.cell(row=1, column=1, value="投递统计看板").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"更新时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws2.cell(row=4, column=1, value=f"总投递数: {total}").font = Font(bold=True, size=12)

    # 状态分布
    ws2.cell(row=6, column=1, value="按状态分布").font = Font(bold=True, size=12)
    ws2.cell(row=7, column=1, value="状态").font = Font(bold=True)
    ws2.cell(row=7, column=2, value="数量").font = Font(bold=True)
    ws2.cell(row=7, column=3, value="占比").font = Font(bold=True)
    for i, (s, n) in enumerate(by_status, 8):
        ws2.cell(row=i, column=1, value=s)
        ws2.cell(row=i, column=2, value=n)
        ws2.cell(row=i, column=3, value=f"{n/total*100:.1f}%" if total else "0%")

    # 渠道分布
    start = 8 + len(by_status) + 2
    ws2.cell(row=start, column=1, value="按渠道分布").font = Font(bold=True, size=12)
    ws2.cell(row=start+1, column=1, value="渠道").font = Font(bold=True)
    ws2.cell(row=start+1, column=2, value="数量").font = Font(bold=True)
    for i, (c, n) in enumerate(by_channel, start+2):
        ws2.cell(row=i, column=1, value=c)
        ws2.cell(row=i, column=2, value=n)

    # 公司分布
    start2 = start + len(by_channel) + 4
    ws2.cell(row=start2, column=1, value="投递公司一览").font = Font(bold=True, size=12)
    ws2.cell(row=start2+1, column=1, value="公司").font = Font(bold=True)
    ws2.cell(row=start2+1, column=2, value="投递数").font = Font(bold=True)
    for i, (c, n) in enumerate(by_company, start2+2):
        ws2.cell(row=i, column=1, value=c)
        ws2.cell(row=i, column=2, value=n)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    # 保存
    export_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, "投递追踪表.xlsx")
    wb.save(filepath)
    print(f"\n  📊 Excel 已导出: {filepath}")
    print(f"  共 {len(rows)} 条投递记录")
    return filepath


def _print_records(rows):
    """命令行打印记录"""
    if not rows:
        print("\n  暂无投递记录。用 'python tracker.py add' 添加第一条吧。")
        return
    print(f"\n{'='*100}")
    print(f"投递记录 ｜ 共 {len(rows)} 条")
    print(f"{'='*100}")
    for r in rows:
        print(f"\n  #{r[0]} [{r[5]}] {r[1]} - {r[2]}")
        print(f"     投递: {r[3]}  渠道: {r[4]}  更新: {r[9]}")
        if r[6]:
            print(f"     反馈: {r[6]}")
        if r[7]:
            print(f"     下一步: {r[7]}")
        if r[10]:
            print(f"     备注: {r[10]}")


# ==================== 交互式命令行 ====================

def interactive_add():
    """交互式添加投递记录"""
    print("\n📝 添加投递记录")
    print("-" * 30)
    company = input("公司名称: ").strip()
    if not company:
        print("公司名不能为空")
        return
    position = input("岗位名称: ").strip()
    if not position:
        print("岗位名不能为空")
        return
    print(f"投递渠道 ({'/'.join(CHANNELS)}) [默认: 官网]: ", end="")
    channel = input().strip() or "官网"
    print(f"当前状态 ({'/'.join(STATUS_NAMES)}) [默认: 已投递]: ", end="")
    status = input().strip() or "已投递"
    if status not in STATUS_NAMES:
        print(f"  未知状态 '{status}'，已设为'已投递'")
        status = "已投递"
    feedback = input("反馈（收到笔试/面试通知等，可留空）: ").strip()
    next_step = input("下一步（准备笔试/等通知等，可留空）: ").strip()
    remark = input("备注（可留空）: ").strip()
    add_record(company, position, channel, status, feedback, next_step, "", remark)


def interactive_update():
    """交互式更新记录"""
    rows = list_records()
    if not rows:
        return
    print()
    rid = input("输入要更新的记录序号 (#): ").strip()
    if not rid.isdigit():
        print("序号必须是数字")
        return
    rid = int(rid)
    conn = _get_conn()
    row = conn.execute("SELECT * FROM applications WHERE id=?", (rid,)).fetchone()
    conn.close()
    if not row:
        print(f"找不到记录 #{rid}")
        return

    print(f"\n当前记录: #{row[0]} [{row[5]}] {row[1]} - {row[2]}")
    print(f"当前反馈: {row[6] or '无'}")
    print(f"当前下一步: {row[7] or '无'}")
    print()
    print(f"新状态 ({'/'.join(STATUS_NAMES)}) [回车保持不变]: ", end="")
    status = input().strip() or None
    feedback = input("新反馈 [回车保持不变]: ").strip()
    feedback = feedback if feedback else None
    next_step = input("新下一步 [回车保持不变]: ").strip()
    next_step = next_step if next_step else None
    remark = input("新备注 [回车保持不变]: ").strip()
    remark = remark if remark else None
    update_record(rid, status, feedback, next_step, remark)


def interactive_main():
    """命令行入口"""
    if len(sys.argv) < 2:
        print(__doc__)
        return
    cmd = sys.argv[1]
    if cmd == "add":
        interactive_add()
    elif cmd == "update":
        interactive_update()
    elif cmd == "list":
        sf = sys.argv[2] if len(sys.argv) > 2 else None
        list_records(sf)
    elif cmd == "export":
        export_excel()
    elif cmd == "stats":
        get_stats()
    else:
        print(f"未知命令: {cmd}")
        print(__doc__)


if __name__ == "__main__":
    interactive_main()
